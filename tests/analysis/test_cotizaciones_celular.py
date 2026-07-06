"""Tests for cotizaciones Excel parsing and cellphone merge helpers."""

import pytest

from business_analyzer.analysis.cotizaciones_celular import (
    DEFAULT_HEADER_ROW_INDEX,
    OUTPUT_COLUMNS,
    admterceros_qualified_table,
    build_column_index_map,
    build_terceros_celular_query,
    distinct_normalized_cedulas,
    enrichment_summary,
    export_enriched_xlsx,
    find_header_row_index,
    format_enriched_rows,
    merge_celular_lookup,
    normalize_cedula,
    parse_cotizaciones_rows,
)


def _crystal_fixture_rows() -> list:
    """Minimal Crystal Reports layout: banner + header + section + CT rows."""
    rows = [[] for _ in range(DEFAULT_HEADER_ROW_INDEX)]
    rows.append(
        [
            "",
            "Docto",
            "",
            "",
            "Numero",
            "",
            "",
            "",
            "Fecha",
            "",
            "Cedula",
            "",
            "",
            "Nombres",
            "",
            "",
            "",
            "",
            "",
            "Detalle",
        ]
    )
    rows.append(["", "COTIZACIONES FACTURADAS"])
    rows.append(
        [
            "",
            "CT",
            "",
            "",
            75776.0,
            "",
            "",
            "",
            "16/06/2026",
            "",
            1080933848.0,
            "",
            "",
            "JUAN PEREZ",
            "",
            "",
            "",
            "",
            "",
            "Tornillo",
        ]
    )
    rows.append(
        [
            "",
            "CT",
            "",
            "",
            75069.0,
            "",
            "",
            "",
            "05/06/2026",
            "",
            "901520709",
            "",
            "",
            "DEPOSITO TRUJILLO",
            "",
            "",
            "",
            "",
            "",
            "Cemento",
        ]
    )
    rows.append(
        [
            "",
            "FV",
            "",
            "",
            12345.0,
            "",
            "",
            "",
            "01/06/2026",
            "",
            "999",
            "",
            "",
            "FACTURA",
            "",
            "",
            "",
            "",
            "",
            "Item",
        ]
    )
    return rows


class TestNormalizeCedula:
    def test_float_integer(self):
        assert normalize_cedula(1080933848.0) == "1080933848"

    def test_scientific_string(self):
        assert normalize_cedula("1.080933848e9") == "1080933848"

    def test_string_with_decimal_suffix(self):
        assert normalize_cedula("36303111.0") == "36303111"

    def test_empty_and_nan(self):
        assert normalize_cedula("") == ""
        assert normalize_cedula(None) == ""
        assert normalize_cedula(float("nan")) == ""


class TestParseCotizacionesRows:
    def test_finds_header_and_filters_ct_only(self):
        rows = _crystal_fixture_rows()
        assert find_header_row_index(rows) == DEFAULT_HEADER_ROW_INDEX

        parsed = parse_cotizaciones_rows(rows)
        assert len(parsed) == 2
        assert all(row["Docto"] == "CT" for row in parsed)

    def test_normalizes_cedula_on_parse(self):
        parsed = parse_cotizaciones_rows(_crystal_fixture_rows())
        assert parsed[0]["CedulaNormalized"] == "1080933848"
        assert parsed[1]["CedulaNormalized"] == "901520709"

    def test_build_column_map_requires_cedula(self):
        with pytest.raises(ValueError, match="Missing required columns"):
            build_column_index_map(["Docto", "Numero"])


class TestAdmTercerosQuery:
    def test_qualified_table_name(self):
        assert admterceros_qualified_table() == "J3System.dbo.AdmTerceros"

    def test_batch_query_uses_qualified_table(self):
        query = build_terceros_celular_query(3)
        assert "FROM J3System.dbo.AdmTerceros" in query
        assert query.count("%s") == 3


class TestMergeCelularLookup:
    def test_merge_attaches_celular_and_match_flag(self):
        rows = parse_cotizaciones_rows(_crystal_fixture_rows())
        lookup = {"1080933848": "3125176217", "901520709": ""}
        enriched = merge_celular_lookup(rows, lookup)

        assert list(enriched[0].keys()) == list(OUTPUT_COLUMNS)
        assert enriched[0]["TercerosCelular"] == "3125176217"
        assert enriched[0]["CelularMatch"] is True
        assert enriched[1]["TercerosCelular"] == ""
        assert enriched[1]["CelularMatch"] is True

    def test_unmatched_cedula_has_no_match_flag(self):
        rows = [{"CedulaNormalized": "000", "Nombres": "X"}]
        enriched = merge_celular_lookup(rows, {})
        assert enriched[0]["TercerosCelular"] == ""
        assert enriched[0]["CelularMatch"] is False


class TestExportEnrichedXlsx:
    def test_empty_celular_written_as_blank_string(self, tmp_path):
        rows = [
            {
                "Docto": "CT",
                "Numero": 1,
                "Fecha": "01/06/2026",
                "Cedula": "901520709",
                "CedulaNormalized": "901520709",
                "Nombres": "TEST",
                "Detalle": "Item",
                "TercerosCelular": "",
                "CelularMatch": True,
            }
        ]
        output = tmp_path / "out.xlsx"
        export_enriched_xlsx(rows, str(output))

        import pandas as pd
        from openpyxl import load_workbook

        frame = pd.read_excel(output, keep_default_na=False, dtype=str)
        assert frame.loc[0, "TercerosCelular"] == ""

        workbook = load_workbook(output)
        worksheet = workbook.active
        celular_col = list(OUTPUT_COLUMNS).index("TercerosCelular") + 1
        assert worksheet.cell(row=2, column=celular_col).value == ""


class TestEnrichmentSummary:
    def test_summary_counts(self):
        rows = parse_cotizaciones_rows(_crystal_fixture_rows())
        lookup = {"1080933848": "3125176217", "901520709": ""}
        enriched = merge_celular_lookup(rows, lookup)
        summary = enrichment_summary(rows, lookup, enriched_rows=enriched)

        assert summary["total_rows"] == 2
        assert summary["distinct_cedulas"] == 2
        assert summary["admterceros_matches"] == 2
        assert summary["matched_cedulas_populated"] == 2
        assert summary["non_empty_celular"] == 1
        assert summary["sample_cedula"] == "1080933848"
        assert summary["sample_celular"] == "3125176217"

    def test_distinct_cedulas_sorted_unique(self):
        rows = [
            {"CedulaNormalized": "2"},
            {"CedulaNormalized": "1"},
            {"CedulaNormalized": "2"},
            {"CedulaNormalized": ""},
        ]
        assert distinct_normalized_cedulas(rows) == ["1", "2"]
